import os
import json
from openpyxl import load_workbook
from gtts import gTTS
import logging
import re
import time

# Dosya yolları
kelimeler_dosyasi = 'kelimeler.xlsx'
sesler_klasoru = 'sesler'

log_klasoru = 'log'

# Log dosyası için yapılandırma
if not os.path.exists(log_klasoru):
    os.makedirs(log_klasoru)

logging.basicConfig(filename=os.path.join(log_klasoru, 'hata_log.txt'), 
                    level=logging.ERROR, 
                    format='%(asctime)s - %(levelname)s - %(message)s')

# Ses dosyaları için klasör oluştur
if not os.path.exists(sesler_klasoru):
    os.makedirs(sesler_klasoru)

# Excel dosyasını oku
def kelimeleri_oku():
    wb = load_workbook(kelimeler_dosyasi, data_only=True)  # data_only=True, formüller yerine hesaplanan değerleri alır
    sheet = wb.active
    kelimeler = []
    
    # Excel dosyasındaki her satırı oku
    for row in sheet.iter_rows(min_row=2, values_only=True):
        arabic_word, turkish_meaning = row
        
        # Arapça kelime veya Türkçe anlam boşsa ya da ArrayFormula varsa atla
        if not arabic_word or not turkish_meaning or isinstance(arabic_word, str) and arabic_word.startswith('='):
            logging.error(f"Geçersiz veya formül içeren satır bulundu. Satır: {row}")
            continue
        
        kelimeler.append((arabic_word, turkish_meaning))
    
    return kelimeler

# Arapça kelimeleri ses dosyasına dönüştür
def kelimeleri_seslendir(kelimeler):
    ses_dosyalar = {}
    
    for arabic_word, turkish_meaning in kelimeler:
        try:
            # Dosya adı için güvenli hale getir
            arabic_word_safe = re.sub(r'[^\w\s-]', '', arabic_word).strip()
            ses_dosyasi = os.path.join(sesler_klasoru, f"{arabic_word_safe}.mp3")
            
            # Kelimeyi seslendir
            tts = gTTS(text=arabic_word, lang='ar')
            tts.save(ses_dosyasi)
            
            # JSON dosyasına kaydetmek için ses dosyasının yolunu sakla
            ses_dosyalar[arabic_word] = {
                'turkish_meaning': turkish_meaning,
                'sound_url': ses_dosyasi
            }
            
            # Bir sonraki isteğe geçmeden önce bekleme süresi ekle (örn. 1 saniye)
            time.sleep(1)  # Bekleme süresi, API sınırlarına ulaşmamak için
        except Exception as e:
            logging.error(f"{arabic_word} kelimesi için ses dosyası oluşturulamadı. Hata: {e}")
    
    return ses_dosyalar

# JSON dosyasına ses dosyası bilgilerini kaydet (benzersiz olacak şekilde)
def ses_wav_json_olustur(ses_dosyalar):
    if os.path.exists(ses_wav_json):
        # JSON dosyasını oku
        with open(ses_wav_json, 'r', encoding='utf-8') as f:
            existing_data = json.load(f)
    else:
        existing_data = {}

    # Yeni verileri mevcut verilerle birleştir
    for arabic_word, data in ses_dosyalar.items():
        # Eğer kelime zaten varsa, atla
        if arabic_word in existing_data:
            logging.info(f"{arabic_word} zaten mevcut, atlanıyor.")
            continue
        
        # Eğer kelime yoksa, ekle
        existing_data[arabic_word] = data

    # Güncellenmiş JSON dosyasını kaydet
    with open(ses_wav_json, 'w', encoding='utf-8') as f:
        json.dump(existing_data, f, ensure_ascii=False, indent=4)

def main():
    try:
        # Kelimeleri oku
        kelimeler = kelimeleri_oku()
        
        # Kelimeleri seslendir
        ses_dosyalar = kelimeleri_seslendir(kelimeler)
        
        # JSON dosyasına kaydet
        ses_wav_json_olustur(ses_dosyalar)
        
    except Exception as e:
        logging.error(f"Main fonksiyonunda hata: {e}")

if __name__ == '__main__':
    main()
